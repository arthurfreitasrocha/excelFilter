from openpyxl.workbook.workbook import Workbook

class Planilha:

    def __init__(self, planilhaForms) -> None:

        # Conta a quantidade de planilhas
        self.contaPlanilha = 0

        # Cria uma nova pasta de trabalho
        novaPastaTrabalho = Workbook()

        self.novaPastaTrabalho = novaPastaTrabalho
        self.planilhaForms = planilhaForms


    def __escrevePlanilha(self, planilha, dados, celulas):

        """
        Utiliza os parâmetros acima para escrever numa "planilha"
        os "dados" capturados nas "celulas" informadas.
        """


        contadores = {
            "contaGrupo": 0,
            "contaCelula": 0
        }

        for chave in dados:

            for informacao in dados[chave]:

                grupo = celulas[contadores["contaGrupo"]] # Captura a célula presente na lista "celulas" de acordo com o índice do contador
                celula = grupo[contadores["contaCelula"]]
                planilha[celula] = informacao # Escreve na planilha a informação presente no dicionário informado

                contadores["contaCelula"] += 1

            contadores["contaCelula"] = 0
            contadores["contaGrupo"] += 1


    def criaPlanilhaDependente(self, nomePlanilha, intervaloLeitura, intervaloEscrita, sequenciaChaves):

        """
        Cria uma planilha com as informações da planilha passada como parâmetro
        """

        # Variável base
        planilhaForms = self.planilhaForms


        # Se nenhuma planilha foi criada, então a primeira é feita.
        if self.contaPlanilha == 0:
            novaPlanilha = self.novaPastaTrabalho.active
            novaPlanilha.title = nomePlanilha
            self.contaPlanilha += 1

        else:
            novaPlanilha = self.novaPastaTrabalho.create_sheet(nomePlanilha)


        # O objeto "leCelulas" será responsável pela leitura dos dados da célula
        intervalo = intervaloLeitura
        leCelulas = Celulas(planilhaForms, intervalo, sequenciaChaves)

        dados = leCelulas.getInformacoes() # Retorna um dicionário com as informações dos usuários


        # O objeto "escreveCelulas" irá capturar as células onde os dados serão escritos
        intervalo = intervaloEscrita
        escreveCelulas = Celulas(planilhaForms, intervalo, sequenciaChaves)

        celulas = escreveCelulas.getCelulas() # Retorna as células de acordo com o intervalo passado

        self.__escrevePlanilha(novaPlanilha, dados, celulas)


        return self.novaPastaTrabalho


    def criaPlanilhaIndependente(self, nomePlanilha, dados, intervaloEscrita, sequenciaChaves, quantLinhas):

        """
        Cria uma planilha com os "dados" passados pelo usuário através de um dicionário
        """

        # Variável base
        planilhaForms = self.planilhaForms


        # Se nenhuma planilha foi criada, então a primeira é feita.
        if self.contaPlanilha == 0:
            novaPlanilha = self.novaPastaTrabalho.active
            novaPlanilha.title = nomePlanilha
            self.contaPlanilha += 1

        else:
            novaPlanilha = self.novaPastaTrabalho.create_sheet(nomePlanilha)


        # O objeto "escreveCelulas" irá capturar as células onde os dados serão escritos
        intervalo = intervaloEscrita
        escreveCelulas = Celulas(planilhaForms, intervalo, sequenciaChaves)

        celulas = escreveCelulas.getCelulas(quantLinhas=quantLinhas) # Retorna as células de acordo com o intervalo passado

        self.__escrevePlanilha(novaPlanilha, dados, celulas)

        return self.novaPastaTrabalho


class Celulas:

    def __init__(self, planilhaForms, intervalo, sequenciaChaves) -> None:

        self.planilhaForms = planilhaForms
        self.intervalo = intervalo
        self.sequenciaChaves = sequenciaChaves


    def __contaUsuarios(self) -> int:

        """
        Recebe como parâmetro uma planilha do Forms e idêntifica quantas pessoas responderam o mesmo.
        """

        planilhaForms = self.planilhaForms

        contadores = {
            'nCelula': 2,
            'quantUsuarios': 0
        }

        while(True):

            if contadores['quantUsuarios'] == 0:
                celula = "A2"

            if planilhaForms[celula].value == None:
                break

            else:
                contadores['nCelula'] += 1
                contadores['quantUsuarios'] += 1
                celula = f"A{contadores['nCelula']}"


        return contadores['quantUsuarios']


    def __geraDicionario(self) -> dict:

            """
            Usando uma sequência de chaves passadas como parâmetro, a função gera um dicionário no formato
            necessário para se adicionar as informações nas planilhas.
            """

            # Cria uma lista com as chaves separadas
            chavesSeparadas = self.sequenciaChaves.split(":")

            # Gera um dicionário vazio com uma lista para cada chave.
            dicionarioChaves = {}
            for chave in chavesSeparadas:
                dicionarioChaves[chave] = []

            return dicionarioChaves


    def __somaColuna(self, sequenciaColuna: list) -> str:

        """
        Retorna a próxima coluna do Excel com base na coluna recebida como parâmetro
        """

        # Joga numa lista os caracteres da sequência e conta quantos "Z" existem na sequência...
        listSequenciaColuna = []
        contaZ = 0
        tamSequencia = len(sequenciaColuna)
        for letra in sequenciaColuna:
            if letra == "Z":
                contaZ += 1

            listSequenciaColuna.append(letra)


        # ... Se toda a sequência for composta por "Z", então a função alterar todas as letras para "A" e adiciona uma letra "A";
        if contaZ == tamSequencia:

            listSequenciaColuna = []

            i = 0
            while(i < tamSequencia+1):
                listSequenciaColuna.append("A")
                i += 1


            # O laço abaixo pega a sequência alfabética inserida na lista (ou seja, a sequencia informada no parâmetro +1)
            # e a insere na variável "sequenciaColuna".
            sequenciaColuna = ""
            for letra in listSequenciaColuna:
                sequenciaColuna += letra

            return sequenciaColuna


        # Do contrário, a função soma normalmente uma letra à coluna.
        # A soma é feita através da conversão da letra para um valor correspondente da tabela ASCII e adicionado +1.
        # Feita a soma, o valor é convertido novamente para String.
        else:

            # O laço de repetição soma + 1 à coluna informada
            listSequenciaColuna = []
            contadores = {
                "contaLaco": tamSequencia - 1,
                "contaLetra": 0
            }
            while(contadores["contaLaco"] >= 0):

                letra = sequenciaColuna[ contadores["contaLaco"] ]

                # Na primeira iteração do laço, a letra é somada normalmente
                if contadores["contaLaco"] + 1 == len(sequenciaColuna):
                    temp = ord(letra) + 1

                # A partir da segunda iteração é verificado se a letra anterior é > "Z".
                # Se sim, então a letra anterior vira "A" e soma-se 1 a letra atual.
                else:
                    if ord(listSequenciaColuna[ contadores["contaLetra"] ]) == 91:
                        listSequenciaColuna[ contadores["contaLetra"] ] = "A"
                        temp = ord(letra) + 1
                        contadores["contaLetra"] += 1

                    else:
                        temp = ord(letra)

                temp = chr(temp)
                listSequenciaColuna.insert(0, temp)

                contadores["contaLaco"] -= 1


            # O laço abaixo pega a sequência alfabética inserida na lista (ou seja, a sequencia informada no parâmetro +1)
            # e a insere na variável "sequenciaColuna".
            sequenciaColuna = ""
            for letra in listSequenciaColuna:
                sequenciaColuna += letra

            return sequenciaColuna


    def __geraSequenciaAlfabetica(self) -> list:

        """
        Usando um intervalo passado como parâmetro, a função retorna as colunas do Excel
        presentes entre o início e o fim desse intervalo.
        """


        # Variáveis para armazenar o intervalo informado
        intervalo = self.intervalo.split(":")

        if len(intervalo) == 1:
            inicioIntervalo = fimIntervalo = intervalo[0]

        elif len(intervalo) == 2:
            inicioIntervalo = intervalo[0]
            fimIntervalo = intervalo[1]

        else:
            colunasUnicas = []
            inicioIntervalo = intervalo[-2]
            fimIntervalo = intervalo[-1]

            for coluna in intervalo:
                if coluna == inicioIntervalo:
                    break

                colunasUnicas.append(coluna)


        # Laço de repetição pra colocar as letras do intervalo numa lista
        letrasSeparadas = []
        for letra in inicioIntervalo:
            letrasSeparadas.append(letra)


        # Cria uma lista com as letras presentes do começo ao fim do intervalo passado
        listaIntervalo = []
        if len(intervalo) > 2:
            for colunas in colunasUnicas: listaIntervalo.append(colunas)
            listaIntervalo.append(inicioIntervalo)

        else:
            listaIntervalo.append(inicioIntervalo)


        if len(intervalo) > 1:

            while(True):

                # A variável "temp" é usada para capturar a próxima coluna retornada pela funcao "somaColuna"
                temp = self.__somaColuna(letrasSeparadas)


                # Feito isso é realizada uma validação, se a coluna armazenada em "temp" for a mesma
                # presente na variável "fimIntervalo", então o laço já capturou todas as colunas presentes
                # do início ao fim do intervalo, encerrando assim o laço.
                if temp == fimIntervalo:
                    listaIntervalo.append(temp)
                    break
                else:
                    listaIntervalo.append(temp)


                # Se a condição à cima retornar falso, então é necessário criar uma nova lista com a coluna armazenada em temp,
                # para ser passada como parâmetro novamente na função "somaColuna".
                letrasSeparadas = []
                for letra in temp:
                    letrasSeparadas.append(letra)


        return listaIntervalo


    def __geraCelulas(self, **kws) -> list:

        """
        A partir de um intervalo, retorna uma lista aninhada com as células desse intervalo.
        Dependendo da quantidade de usuários que responderam ao Forms, podem haver mais ou menos células.
        """

        quantLinhas = kws.get("quantLinhas")

        # Variáveis base
        if quantLinhas == None:
            quantUsuarios = self.__contaUsuarios() + 1
        else:
            quantUsuarios = quantLinhas

        sequenciaAlfabetica = self.__geraSequenciaAlfabetica()

        # Cria laços de repetição aninhados.
        # Um para percorrer a "sequenciaAlfabetica" e o outro para adicionar o valor da linha à cada coluna da sequência,
        # formando assim as células.
        listaCelulas = []
        for coluna in sequenciaAlfabetica:

            temp = []
            contaLaco = 0
            while(contaLaco < quantUsuarios):

                sequenciaCelula = f"{coluna}{contaLaco+1}"
                temp.append(sequenciaCelula)

                contaLaco += 1

            listaCelulas.append(temp)


        return listaCelulas



    def getInformacoes(self) -> dict:

        """
        Utiliza o intervalo informado por parâmetro para buscar as informações
        na planilha passada como parâmetro do objeto e retornar um dicionário
        contendo as mesmas.
        """

        planilhaForms = self.planilhaForms

        # Gera um dicionário com listas vazias
        dicionarioChaves = self.__geraDicionario()
        dicionario = dicionarioChaves
        chaves = self.sequenciaChaves.split(":")

        # Gera uma lista com as células que serão varridas
        celulas = self.__geraCelulas()

        # Faz uma varredura nas células selecionadas e armazena as informações no dicionário
        indiceChave = 0
        for grupo in celulas:

            for celula in grupo:
                temp = chaves[indiceChave]
                dicionario[temp].append(planilhaForms[celula].value)

            indiceChave += 1


        return dicionario


    def getCelulas(self, **kws) -> list:

        """
        Retorna as células correspondentes ao intervalo passado para serem
        usadas na inserção de conteúdo da planilha.
        """

        quantLinhas = kws.get("quantLinhas")

        return self.__geraCelulas(quantLinhas=quantLinhas)


    def getSomaColuna(self, sequenciaColuna: list) -> str:

        """
        Retorna para o usuário a sua "sequenciaColuna" com a adição de 1 coluna.
        """

        return self.__somaColuna(sequenciaColuna)

